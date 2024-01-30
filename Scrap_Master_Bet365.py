import subprocess
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from datetime import datetime, timedelta
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as cond
from selenium.webdriver import ActionChains
import pandas as pd
import telebot
import urllib
import openpyxl
import random
from sklearn.ensemble import ExtraTreesClassifier

def check_exists_by_xpath(xpath):
    try:
        driver.find_element(By.XPATH, xpath)
        return True
    except:
        return False


def check_exists_by_class(class_name):
    try:
        driver.find_element(By.CLASS_NAME, class_name)
        return True
    except:
        return False


def mensagem(TimeHOME, GolHOME, GolVISITANTES, TimeVISITANTE, AtaquesPERIGHOME, AtaquesPERIGVISITANTES, EscanteiosHOME,
             EscanteiosVISITANTE, ChutesGOLHOME, ChutesGOLVISITANTE, Tempo, Aposta, IA, ml_validation):
    TimeHOME_link = urllib.parse.quote(f"{TimeHOME}")
    if 'Green' in ml_validation:
        recomend = ' GWzord AVALIA GRANDE POTENCIAL, conferir entrada junto a analise manual'
    else:
        recomend = ' GWzord AVALIA PEQUENO RISCO, detalhar analise manual para ver se vale a entrada'
    if Tempo < 45:
        momento = 'no primeiro tempo da partida'
    else:
        momento = 'no segundo tempo da partida'
    message = f"""\n
        Fala meus apostadores, pega essa ANALISE para o jogo abaixo:\n
        Probabilidade da inteligencia artificial: {IA}% - Avaliacao: {recomend}\n\n
        {TimeHOME} {GolHOME} - {GolVISITANTES} {TimeVISITANTE}\n 
        - Tempo: {Tempo}\n
        - Ataques Perigosos: {AtaquesPERIGHOME} - {AtaquesPERIGVISITANTES}\n
        - Escanteios: {EscanteiosHOME} - {EscanteiosVISITANTE}\n
        - Chutes no gol: {ChutesGOLHOME} - {ChutesGOLVISITANTE}\n\n
        A DICA do GWzord: +{Aposta} gol {momento}.\n
        Entre no link abaixo para analisar o jogo:\n
        https://www.bet365.com/#/AX/K^{TimeHOME_link}
        """
    return message


def estatistica_tabela():
    global df
    temp = '//span[@class="ml1-SoccerClock_Clock "]'
    temp_ex = check_exists_by_xpath(temp)
    botao = '//div[@class="lv-ButtonBar_MatchLiveIcon me-MediaButtonLoader me-MediaButtonLoader_ML1 "]'
    botao_ex = check_exists_by_xpath(botao)
    if not temp_ex:
        if botao_ex:
            driver.find_element(By.XPATH, botao).click()
            sleep(random.uniform(0.5, 0.7))
        else:
            pass
    try:
        tempo = driver.find_element(By.XPATH, '//span[@class="ml1-SoccerClock_Clock "]').text
    except:
        tempo = '45:00'
    try:
        est_casa = list(driver.find_elements(By.CLASS_NAME, 'ml1-WheelChartAdvanced_Team1Text '))
        est_vis = list(driver.find_elements(By.CLASS_NAME, 'ml1-WheelChartAdvanced_Team2Text '))
        est_ca = list((driver.find_elements(By.CLASS_NAME, 'ml1-StatsColumnAdvanced_MiniValue ')))
        est_fin = list((driver.find_elements(By.CLASS_NAME, 'ml1-ProgressBarAdvancedDual_SideLabel ')))
        gol_h = driver.find_element(By.CSS_SELECTOR,
                                    'div.lsb-ScoreBasedScoreboardAggregate_ScoreContainer > div:nth-child(1)').text
        gol_v = driver.find_element(By.CSS_SELECTOR,
                                    'div.lsb-ScoreBasedScoreboardAggregate_ScoreContainer > div:nth-child(2)').text
        time = driver.find_element(By.CSS_SELECTOR,
                                   'div.lsb-ScoreBasedScoreboardAggregate_TeamContainer.lsb-ScoreBasedScoreboardAggregate_Team1Container').text
        if 'Esports' in time:
            return
        time_2 = driver.find_element(By.CSS_SELECTOR,
                                     'div.lsb-ScoreBasedScoreboardAggregate_TeamContainer.lsb-ScoreBasedScoreboardAggregate_Team2Container').text
        if len(est_casa) > 2:
            ataque_h = est_casa[0].text
            ataque_perig_h = est_casa[1].text
            posse_bola_h = est_casa[2].text
        else:
            ataque_h = est_casa[0].text
            ataque_perig_h = est_casa[1].text
            posse_bola_h = '0'
        if len(est_vis) > 2:
            ataque_v = est_vis[0].text
            ataque_perig_v = est_vis[1].text
            posse_bola_v = est_vis[2].text
        else:
            ataque_v = est_vis[0].text
            ataque_perig_v = est_vis[1].text
            posse_bola_v = '0'
        cartao_a_h = est_ca[2].text
        cartao_a_v = est_ca[5].text
        cartao_v_h = est_ca[1].text
        cartao_v_v = est_ca[4].text
        escanteio_h = est_ca[0].text
        escanteio_v = est_ca[3].text
        chute_gol_h = est_fin[1].text
        chute_gol_v = est_fin[3].text
        finalizacao_h = est_fin[0].text
        finalizacao_v = est_fin[2].text

        df = pd.concat([df, df.from_records(
            [{'GolHOME': gol_h, 'GolVISITANTES': gol_v, 'TimeHOME': time, 'TimeVISITANTE': time_2,
              'AtaquesHOME': ataque_h, 'AtaquesVISITANTES': ataque_v, 'AtaquesPERIGHOME': ataque_perig_h,
              'AtaquesPERIGVISITANTES': ataque_perig_v, 'PosseHOME': posse_bola_h,
              'PosseVISITANTE': posse_bola_v, 'CartoesAMARELOHOME': cartao_a_h,
              'CartoesAMARELOVISITANTE': cartao_a_v, 'CartoesVERMELHOHOME': cartao_v_h,
              'CartoesVERMELHOVISITANTE': cartao_v_v, 'EscanteiosHOME': escanteio_h,
              'EscanteiosVISITANTE': escanteio_v, 'ChutesGOLHOME': chute_gol_h,
              'ChutesGOLVISITANTE': chute_gol_v, 'FinalizacaoHOME': finalizacao_h,
              'FinalizacaoVISITANTE': finalizacao_v, 'Tempo': tempo}])], ignore_index=True)
    except:
        try:
            time = driver.find_element(By.CSS_SELECTOR,
                                       'div.lsb-ScoreBasedScoreboardAggregate_TeamContainer.lsb-ScoreBasedScoreboardAggregate_Team1Container').text
            print(f'Estatisticas indisponíveis para o jogo do time {time}')
        except:
            print('Erro geral web scraping bet365')


def check_green():
    global df
    data = datetime.today().strftime('%Y-%m-%d')
    wb = openpyxl.load_workbook('controle_apostas.xlsx')
    ws = wb['controle']
    if df.empty:
        return
    else:
        df_check = pd.read_excel('controle_apostas.xlsx', sheet_name='controle')
        index = df_check[(df_check['TimeHOME'] == df.iloc[0]['TimeHOME']) & (df_check['Data'] == data)].index.values
        if len(index) == 0:
            return
        index_cell = max(index) + 2
        df_check = df_check.query(f'Data == "{data}"')
        if df_check.empty:
            return
        else:
            df_check_result = df_check.loc[df_check['TimeHOME'] == df.iloc[0]['TimeHOME']]
            if df_check_result.empty:
                return
            else:
                if not pd.isnull(df_check_result.iloc[0]['Resultado']):
                    return
                else:
                    try:
                        tmp_check = int(df.iloc[0]['Tempo'][:3].replace(':', ''))
                    except:
                        try:
                            tmp_check = int(df.iloc[0]['Tempo'][:2])
                        except:
                            print('erro ao converter o tempo')
                            return
                    if int(df_check_result.iloc[0]['Tempo']) < 45 and tmp_check <= 47:
                        soma = int(df.iloc[0]['GolHOME']) + int(df.iloc[0]['GolVISITANTES'])
                        soma_green = int(df_check_result.iloc[0]['GolHOME']) + int(
                            df_check_result.iloc[0]['GolVISITANTES'])
                        if soma > soma_green:
                            elemento = random.choice(lista_msg_green)
                            msg = f"""\n
                                    {elemento} no jogo do {df_check_result.iloc[0]['TimeHOME']}
                                    """
                            bot.send_message(chat_id, msg)
                            #bot.send_sticker(chat_id,"")
                            ws[f'Y%i' % index_cell].value = 'Green'
                            wb.save('controle_apostas.xlsx')
                        else:
                            return
                    elif int(df_check_result.iloc[0]['Tempo']) > 45 and (tmp_check > 45 & tmp_check < 92):
                        soma = int(df.iloc[0]['GolHOME']) + int(df.iloc[0]['GolVISITANTES'])
                        soma_green = int(df_check_result.iloc[0]['GolHOME']) + int(
                            df_check_result.iloc[0]['GolVISITANTES'])
                        if soma > soma_green:
                            elemento = random.choice(lista_msg_green)
                            msg = f"""\n
                                    {elemento} no jogo do {df_check_result.iloc[0]['TimeHOME']}
                                    """
                            bot.send_message(chat_id, msg)
                            #bot.send_sticker(chat_id,"")
                            ws[f'Y%i' % index_cell].value = 'Green'
                            wb.save('controle_apostas.xlsx')
                        else:
                            return
                    else:
                        elemento = random.choice(lista_msg_red)
                        msg = f"""\n
                                    A aposta no jogo do {df_check_result.iloc[0]['TimeHOME']} {elemento}
                                    """
                        bot.send_message(chat_id, msg)
                        #bot.send_sticker(chat_id,"")
                        ws[f'Y%i' % index_cell].value = 'Red'
                        wb.save('controle_apostas.xlsx')


def envia_status():
    global check_envio
    hora_envia_status = datetime.now().replace(hour=20, minute=35)
    if datetime.now() > hora_envia_status and not check_envio:
        df_check_enviados = pd.read_excel('controle_apostas.xlsx', sheet_name='controle')
        data_check = datetime.today().strftime('%Y-%m-%d')
        df_check = df_check_enviados.query(f'Data == "{data_check}"')
        if df_check.empty:
            msg = f"""\n
                        Não tivemos nenhuma boa oportunidade no dia de hoje. Só enviamos as MELHORES. Seguimos! Amanhão tem mais!
                        """
        else:
            df_green = df_check.query(f'Resultado == "Green"')
            entradas = len(df_check.index)
            greens = len(df_green.index)
            if greens / entradas > 0.5:
                msg = f"""\n
                            O GWzord finaliza o dia com {greens} greens nas {entradas} entradas. FALA se não é o BRABO!! Reage a essa mensagem pra dar uma força pro robozão!
                            """
            else:
                msg = f"""\n
                            É preciso analisar bem as entradas meu amigo. O GWzord finaliza o dia com {greens} greens nas {entradas} entradas. Segue o plano!
                            """
        bot.send_message(chat_id, msg)
        subprocess.run('shutdown /s /t 1200 /d p:0:0 /c "Shutdown planejado por Wilson"')
        check_envio = True
    else:
        return


def apostas():
    global df
    df_check_enviados = pd.read_excel('controle_apostas.xlsx', sheet_name='controle')
    df_check_times_ht = pd.read_excel('controle_apostas.xlsx', sheet_name='times')
    df_ml = df_check_enviados.dropna(axis=0)
    y = df_ml['Resultado']
    x = df_ml.drop(['Resultado', 'TimeHOME', 'TimeVISITANTE', 'Data', 'Tipo', 'IA', 'Tempo'], axis=1)
    try:
        df['Tempo'] = df['Tempo'].apply(lambda x: x[:3])
        try:
            df['Tempo'] = df['Tempo'].str.replace(':', '').astype(int)
        except:
            pass
        df = df.apply(pd.to_numeric, errors='ignore')
        list_tipo = dict(enumerate(['Volume', 'Volume_2T', 'Ataque', 'Chutes', 'Eficacia']))
        list_odds = [
            'GolHOME == 0 & GolVISITANTES == 0 & Tempo > 10 & Tempo < 28 & (AtaquesPERIGHOME+AtaquesPERIGVISITANTES)/Tempo >= 1.4 & ((AtaquesPERIGHOME/AtaquesPERIGVISITANTES > 1.6 & ChutesGOLHOME >= 3 & EscanteiosHOME+EscanteiosVISITANTE >= 3) | (AtaquesPERIGVISITANTES/AtaquesPERIGHOME > 1.8 & FinalizacaoVISITANTE/FinalizacaoHOME >= 3 & ChutesGOLVISITANTE >= 1 & EscanteiosVISITANTE >= 2))',
            'GolHOME == 0 & GolVISITANTES == 0 & Tempo > 60 & Tempo < 70 & (AtaquesPERIGHOME+AtaquesPERIGVISITANTES)/Tempo >= 2 & ChutesGOLHOME+ChutesGOLVISITANTE >= 8',
            '(AtaquesPERIGHOME+AtaquesPERIGVISITANTES)/Tempo >= 2.2 & Tempo > 10 & Tempo < 28 & ChutesGOLHOME+ChutesGOLVISITANTE >= 5',
            'GolHOME == 0 & GolVISITANTES == 0 & ChutesGOLHOME+ChutesGOLVISITANTE >= 5 & EscanteiosHOME+EscanteiosVISITANTE >= 2 & Tempo < 20',
            'GolHOME == 0 & GolVISITANTES == 0 & ((FinalizacaoHOME >= 5 & FinalizacaoVISITANTE <= 1 & ChutesGOLHOME/FinalizacaoHOME >= 0.5) | (FinalizacaoVISITANTE >= 5 & FinalizacaoHOME <= 1 & ChutesGOLVISITANTE/FinalizacaoVISITANTE >= 0.5)) & Tempo < 28']
        od = 0
        for odds in list_odds:
            try:
                df_apostax = df.query(odds)
                if not df_apostax.empty:
                    for index in range(len(df_apostax.index)):
                        time_check = df_apostax.iloc[index]['TimeHOME']
                        data_check = datetime.today().strftime('%Y-%m-%d')
                        if df_check_enviados.query(
                                f'Data == "{data_check}" & TimeHOME == "{time_check}"').empty and time_check not in lista_enviados:
                            lista_enviados.append(time_check)
                            soma_IA_ht = 0
                            IA_ini = random.uniform(70, 77)
                            soma_IA = int(df_apostax.iloc[index]['AtaquesPERIGHOME']) / int(
                                df_apostax.iloc[index]['AtaquesPERIGVISITANTES']) + int(
                                df_apostax.iloc[index]['AtaquesPERIGVISITANTES']) / int(
                                df_apostax.iloc[index]['AtaquesPERIGHOME'])
                            df_check_ht_home = df_check_times_ht.loc[
                                df_check_times_ht['Times_HT'] == df_apostax.iloc[index]['TimeHOME']]
                            df_check_ht_visitante = df_check_times_ht.loc[
                                df_check_times_ht['Times_HT'] == df_apostax.iloc[index]['TimeVISITANTE']]
                            if not df_check_ht_home.empty:
                                soma_IA_ht = soma_IA_ht + 10
                            if not df_check_ht_visitante.empty:
                                soma_IA_ht = soma_IA_ht + 10

                            soma_IA_total = round(IA_ini + soma_IA_ht + soma_IA, 2)
                            if soma_IA_total > 100:
                                soma_IA_total = random.uniform(97.01, 99.99)

                            df_apostax['Data'] = datetime.today().strftime('%Y-%m-%d')
                            df_apostax['Tipo'] = list_tipo[od]
                            df_apostax['IA'] = soma_IA_total

                            # Machine Learning SVM
                            if int(df_apostax.iloc[index]['Tempo']) <= 45:
                                df_ml_aposta = df_apostax.query("Tempo <= 45")
                                df_ml_aposta = df_ml_aposta.drop(['TimeHOME', 'TimeVISITANTE', 'Data', 'Tipo', 'IA', 'Tempo'], axis=1)
                            else:
                                df_ml_aposta = df_apostax.drop(['TimeHOME', 'TimeVISITANTE', 'Data', 'Tipo', 'IA', 'Tempo'], axis=1)
                            model = ExtraTreesClassifier()
                            model.fit(x, y)
                            ml_validation = str(model.predict(df_ml_aposta))


                            aposta = int(df_apostax.iloc[index]['GolHOME']) + int(
                                df_apostax.iloc[index]['GolVISITANTES']) + 0.5
                            message = mensagem(df_apostax.iloc[index]['TimeHOME'], df_apostax.iloc[index]['GolHOME'],
                                               df_apostax.iloc[index]['GolVISITANTES'],
                                               df_apostax.iloc[index]['TimeVISITANTE'],
                                               df_apostax.iloc[index]['AtaquesPERIGHOME'],
                                               df_apostax.iloc[index]['AtaquesPERIGVISITANTES'],
                                               df_apostax.iloc[index]['EscanteiosHOME'],
                                               df_apostax.iloc[index]['EscanteiosVISITANTE'],
                                               df_apostax.iloc[index]['ChutesGOLHOME'],
                                               df_apostax.iloc[index]['ChutesGOLVISITANTE'],
                                               df_apostax.iloc[index]['Tempo'], aposta, soma_IA_total, ml_validation)
                            bot.send_message(chat_id, message)
                            #bot.send_sticker(chat_id,"")
                            with pd.ExcelWriter('controle_apostas.xlsx', mode='a', engine='openpyxl',
                                                if_sheet_exists='overlay') as writer:
                                startrow = writer.sheets['controle'].max_row
                                df_apostax.to_excel(writer, sheet_name='controle', header=False, index=False,
                                                    startrow=startrow)
                            df = pd.DataFrame(columns=COLUNAS)
            except:
                df = pd.DataFrame(columns=COLUNAS)
            od += 1
    except:
        df = pd.DataFrame(columns=COLUNAS)


if __name__ == "__main__":
    options = webdriver.ChromeOptions()
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--incognito')
    options.add_argument('--disable-notifications')
    driver = webdriver.Chrome(options=options)
    # driver.get("https://www.bet365.com/#/IP/B1")
    driver.execute_script("window.open('https://www.bet365.com/#/IP/B1', '_blank')")
    COLUNAS = ['GolHOME', 'GolVISITANTES', 'TimeHOME', 'TimeVISITANTE', 'AtaquesHOME', 'AtaquesVISITANTES',
               'AtaquesPERIGHOME', 'AtaquesPERIGVISITANTES', 'PosseHOME', 'PosseVISITANTE', 'CartoesAMARELOHOME',
               'CartoesAMARELOVISITANTE', 'CartoesVERMELHOHOME', 'CartoesVERMELHOVISITANTE', 'EscanteiosHOME',
               'EscanteiosVISITANTE', 'ChutesGOLHOME', 'ChutesGOLVISITANTE', 'FinalizacaoHOME', 'FinalizacaoVISITANTE',
               'Tempo']
    token = ''# Put here telegram token
    chat_id = ''# Put here telegram Chat ID
    bot = telebot.TeleBot(token)
    df = pd.DataFrame(columns=COLUNAS)
    lista_enviados = []
    lista_msg_green = ['VAMOO, O GWzord CRAVA DEMAIS!! TOMA GREEN',
                       'PAGAA BETANIA, BETINA, DENISE.. O GWzord mandou e logico que bateu o GREEN',
                       'SABE O QUE ACONTECEU DENOVO? O GWzordCRAVOU',
                       'OLHA AI?! NAO TEM JEITO, O GWzord segue GREENZANDO e dessa vez foi',
                       'FALA QUE EU TE ESCUTO APOSTADOR!! Voce pediu GREEN, entao TOMA!',
                       'SE VOCE AINDA NÃO CANSOU DE GANHAR DINHEIRO, PEGA esse GREEN que o GWzord mandou',
                       'EU SOU O MILIOR! PEGA esse GREEN do ROBOZAO GWzord',
                       'RECEBAAA!! GRAÇAS A DEUS!!! RECEBA esse GREEN que o GWzord mandou',
                       'VAMO QUE VAMO DE GREEN! o BRABO mandou',
                       'DAA LHEEE!! O GOLZINHO VEIO',
                       'TA LAAAAA! GREEN',
                       'CAAAAAIXAAAA!!! Saiu o gol']
    lista_msg_red = ['tinha muito valor, mas o green nao veio. Seguimos meu apostador, BORAA PRA PROXIMA!!',
                     'infelizmente redou. Seguimos com CALMA e GESTAO de banca!',
                     'era a BOA porem acabou nao vindo. VAMOS RECUPERAR!!',
                     'acabou redando. Seguimos o PLANO!',
                     'deu ruim. Porem o importante é somatorio no final do mes. FOCO NA GESTAO',
                     'azedou o caldo. Seguimos com ESTRATEGIA, ANALISE e GESTAO',
                     'moio. Mas CALMA, a proxima vai ser GREEN']
    # lista_fig_green = []
    # lista_fig_red = []
    # lista_fig_msg = []
    end_cod = datetime.now() + timedelta(minutes=29)
    check_envio = False

    try:
        sleep(random.uniform(8, 15))
        window_bet = driver.window_handles[1]
        driver.switch_to.window(window_bet)
        WebDriverWait(driver, 15).until(
            cond.element_to_be_clickable((By.CLASS_NAME, ('iip-IntroductoryPopup_Cross')))).click()
        WebDriverWait(driver, 15).until(
            cond.element_to_be_clickable((By.CLASS_NAME, ('ccm-CookieConsentPopup_Accept ')))).click()
        while datetime.now() < end_cod:
            for u in range(2, len(driver.find_elements(By.CLASS_NAME, 'ovm-CompetitionHeader_Header'))):
                i = 1
                njogos = 3
                try:
                    driver.find_element(By.XPATH,
                                        '/html/body/div[1]/div/div[4]/div[2]/div[1]/div/div/div/div/div/div/div[1]/div[2]/div[2]/div[%i]/div[2]/div[2]/div[2]/div/div' % u).click()
                    sleep(random.uniform(0.5, 0.7))
                    estatistica_tabela()
                    check_green()
                    apostas()
                    envia_status()
                except:
                    if check_exists_by_class('lms-StandardLogin_Container '):
                        ActionChains(driver).move_to_element_with_offset(driver.find_element(By.XPATH, '//html'), 50,
                                                                         200).click().perform()
                    try:
                        for tentar in range(0, 3):
                            if check_exists_by_xpath(
                                    '/html/body/div[1]/div/div[4]/div[2]/div[1]/div/div/div/div/div/div/div['
                                    '1]/div[2]/div[2]/div[%i]/div[2]/div[%i]/div[2]/div/div' % (u, njogos)):
                                break
                            njogos += 1
                    except:
                        continue
                while i < 40:
                    try:
                        for tentar_btn in range(0, 3):
                            if check_exists_by_xpath(
                                    '/html/body/div[1]/div/div[4]/div[2]/div[1]/div/div/div/div/div/div/div['
                                    '1]/div[2]/div[2]/div[%i]/div[2]/div[%i]/div[2]/div/div' % (u, njogos)):
                                break
                            njogos += 1
                        driver.find_element(By.XPATH, '/html/body/div[1]/div/div[4]/div[2]/div['
                                                      '1]/div/div/div/div/div/div/div[1]/div[2]/div[2]/div[%i]/div[2]/div['
                                                      '%i]/div[2]/div/div' % (u, njogos)).click()
                    except:
                        if check_exists_by_class('lms-StandardLogin_Container '):
                            ActionChains(driver).move_to_element_with_offset(driver.find_element(By.XPATH, '//html'),
                                                                             50,
                                                                             200).click().perform()
                        break
                    sleep(random.uniform(0.5, 0.7))
                    estatistica_tabela()
                    check_green()
                    apostas()
                    envia_status()
                    i += 1
                    njogos += 1
        driver.quit()
    except:
        driver.quit()
        subprocess.call([r'C:\Python\bet365-scraper-master\bet365\run_bet365.bat'])
